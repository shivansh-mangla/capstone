# preprocess.py
from openpyxl import load_workbook


class PreprocessClass:
    @staticmethod
    def preprocessScriptFunc(path):
        """Load workbook and run processing, returning (finalDict, finalDict2)."""
        workbook = load_workbook(path)
        # debug: print(workbook.sheetnames)
        finalDict, finalDict2 = PreprocessClass.process_workbook(workbook)
        return finalDict, finalDict2

    @staticmethod
    def get_HOUR_position(sheet):
        """Returns the (row, col) of the first place where HOUR/HOURS is written."""
        ans = (0, 0)
        sum1 = 10000
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in ("HOUR", "HOURS"):
                    if cell.row + cell.column < sum1:
                        ans = (cell.row, cell.column)
                        sum1 = cell.row + cell.column
        return ans

    @staticmethod
    def get_DAY_position(sheet):
        """Returns the (row, col) of the last place where DAY/DAYS is written."""
        ans = (0, 0)
        sum1 = -1
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in ("DAY", "DAYS"):
                    if cell.row + cell.column > sum1:
                        ans = (cell.row, cell.column)
                        sum1 = cell.row + cell.column
        return ans

    @staticmethod
    def getCellBorders(row, col, sheet):
        """Returns a string like 'TLRB' denoting which sides of the cell have borders."""
        brdrs = ''
        tmp = sheet.cell(row, col).border
        if tmp.top.style is not None:
            brdrs += 'T'
        if tmp.left.style is not None:
            brdrs += 'L'
        if tmp.right.style is not None:
            brdrs += 'R'
        if tmp.bottom.style is not None:
            brdrs += 'B'
        return brdrs

    @staticmethod
    def is_cell_merged(row, col, merged_ranges):
        """Check if a given cell is part of a merged range.
        Returns (True, range) or (False, None)
        """
        for cell_range in merged_ranges:
            if (
                cell_range.min_row <= row <= cell_range.max_row and
                cell_range.min_col <= col <= cell_range.max_col
            ):
                return (True, cell_range)
        return (False, None)

    @staticmethod
    def create_subgroup_map(sheet):
        """Returns a dict with key=subgroup name, value=start column (assumes width=2 columns)."""
        row_day, end_col = PreprocessClass.get_DAY_position(sheet)
        row_hour, start_col = PreprocessClass.get_HOUR_position(sheet)

        myMap = {}
        col = start_col + 1
        while col < end_col:
            val = sheet.cell(row_hour, col).value
            if val in ('DAY', 'HOURS') or val is None:
                break
            myMap[val] = col
            col += 2
        return myMap

    @staticmethod
    def handle_lecture(row, col, sheet, merged_ranges):
        subject_val = sheet.cell(row, col).value or ""
        subjectCode = subject_val[:-1].strip() if len(subject_val) > 0 else ""
        merged_info = PreprocessClass.is_cell_merged(row, col, merged_ranges)[1]
        min_col, max_col = merged_info.min_col, merged_info.max_col

        HOUR_position_row, HOUR_position_col = PreprocessClass.get_HOUR_position(sheet)
        TIME_position_row = HOUR_position_row + 2
        TIME_position_col = HOUR_position_col

        time1 = sheet.cell(TIME_position_row, TIME_position_col).value
        t1 = row - TIME_position_row
        row += 1
        venue = sheet.cell(row, min_col).value

        check = PreprocessClass.is_cell_merged(row, max_col, merged_ranges)
        if check[0]:
            min_col2 = check[1].min_col
            teacher_code = sheet.cell(row, min_col2).value
        else:
            teacher_code = sheet.cell(row, max_col).value

        while True:
            if 'B' in PreprocessClass.getCellBorders(row, min_col, sheet) or \
               'T' in PreprocessClass.getCellBorders(row + 1, min_col, sheet):
                break
            row += 1

        time2 = sheet.cell(row + 1, TIME_position_col).value
        t2 = row - TIME_position_row

        return {
            'type': 'lecture',
            'subjectCode': subjectCode,
            'venue': venue,
            'teacherCode': teacher_code,
            'time1': time1,
            'time2': time2,
            't1': t1,
            't2': t2
        }

    @staticmethod
    def handle_tutorial(row, col, sheet, merged_ranges):
        subject_val = sheet.cell(row, col).value or ""
        subjectCode = subject_val[:-1].strip() if len(subject_val) > 0 else ""
        merged_info = PreprocessClass.is_cell_merged(row, col, merged_ranges)[1]
        min_col, max_col = merged_info.min_col, merged_info.max_col

        HOUR_position_row, HOUR_position_col = PreprocessClass.get_HOUR_position(sheet)
        TIME_position_row = HOUR_position_row + 2
        TIME_position_col = HOUR_position_col

        time1 = sheet.cell(TIME_position_row, TIME_position_col).value
        t1 = row - TIME_position_row
        row += 1
        venue = sheet.cell(row, min_col).value

        brdr = PreprocessClass.getCellBorders(row, min_col, sheet)
        if 'B' not in brdr:
            row += 1

        check = PreprocessClass.is_cell_merged(row, max_col, merged_ranges)
        if check[0]:
            min_col2 = check[1].min_col
            teacher_code = sheet.cell(row, min_col2).value
        else:
            teacher_code = sheet.cell(row, max_col).value

        while True:
            if 'B' in PreprocessClass.getCellBorders(row, min_col, sheet) or \
               'T' in PreprocessClass.getCellBorders(row + 1, min_col, sheet):
                break
            row += 1

        time2 = sheet.cell(row + 1, TIME_position_col).value
        t2 = row - TIME_position_row

        return {
            'type': 'tutorial',
            'subjectCode': subjectCode,
            'venue': venue,
            'teacherCode': teacher_code,
            'time1': time1,
            'time2': time2,
            't1': t1,
            't2': t2
        }

    @staticmethod
    def handle_lab(row, col, sheet, merged_ranges):
        subject_val = sheet.cell(row, col).value or ""
        subjectCode = subject_val[:-1].strip() if len(subject_val) > 0 else ""
        merged_info = PreprocessClass.is_cell_merged(row, col, merged_ranges)[1]
        min_col, max_col = merged_info.min_col, merged_info.max_col

        HOUR_position_row, HOUR_position_col = PreprocessClass.get_HOUR_position(sheet)
        TIME_position_row = HOUR_position_row + 2
        TIME_position_col = HOUR_position_col

        time1 = sheet.cell(TIME_position_row, TIME_position_col).value
        t1 = row - TIME_position_row
        row += 1
        venue = sheet.cell(row, min_col).value

        row += 1
        teacher_code = sheet.cell(row, max_col).value

        while True:
            if 'B' in PreprocessClass.getCellBorders(row, min_col, sheet) or \
               'T' in PreprocessClass.getCellBorders(row + 1, min_col, sheet):
                break
            row += 1

        time2 = sheet.cell(row + 1, TIME_position_col).value
        t2 = row - TIME_position_row

        return {
            'type': 'lab',
            'subjectCode': subjectCode,
            'venue': venue,
            'teacherCode': teacher_code,
            'time1': time1,
            'time2': time2,
            't1': t1,
            't2': t2
        }

    @staticmethod
    def get_time_table(col, sheet, merged_ranges, hourINC):
        days_of_week = 5
        slots_in_each_day = 14

        HOUR_position_row, HOUR_position_col = PreprocessClass.get_HOUR_position(sheet)
        TIME_position_row = HOUR_position_row + hourINC
        TIME_position_col = HOUR_position_col

        myList = []

        for day in range(days_of_week):
            for slot in range(slots_in_each_day):
                row = TIME_position_row + 2 * (day * slots_in_each_day + slot)

                is_cm = PreprocessClass.is_cell_merged(row, col, merged_ranges)
                if is_cm[0]:
                    temp_col = is_cm[1].min_col  # Lower column of the merged cell range

                    if 'T' in PreprocessClass.getCellBorders(row, temp_col, sheet) or \
                       'B' in PreprocessClass.getCellBorders(row - 1, temp_col, sheet):
                        val = sheet.cell(row, temp_col).value
                        if val is not None:
                            val = str(val).strip()
                            if len(val) > 0:
                                last_char = val[-1]
                                if last_char == 'L':
                                    myList.append(PreprocessClass.handle_lecture(row, temp_col, sheet, merged_ranges))
                                elif last_char == 'P':
                                    myList.append(PreprocessClass.handle_lab(row, temp_col, sheet, merged_ranges))
                                elif last_char == 'T':
                                    myList.append(PreprocessClass.handle_tutorial(row, temp_col, sheet, merged_ranges))
        return myList

    @staticmethod
    def get_time_table2(myMap, sheet, merged_ranges, hrINC):
        myList = {}
        myList2 = {}

        for subgroup in list(myMap.keys()):
            myList[subgroup] = {}

            subgroupDict = myList[subgroup]
            subgroupDict['lecture'] = {}
            subgroupDict['lab'] = {}
            subgroupDict['tutorial'] = {}
            subgroupDict['elective'] = {}

            tt = PreprocessClass.get_time_table(myMap[subgroup], sheet, merged_ranges, hrINC)
            for item in tt:
                sc = item.get('subjectCode', '').strip()
                if '/' not in sc and sc[:6] not in ['UMC743', 'UCS658', 'UCS657', 'UCS748', 'UCS539']:
                    # not an elective course
                    myList2.setdefault(item['subjectCode'], []).append(subgroup)
                    subgroupDict[item['type']][item['subjectCode']] = {'slots': [], 'teacher_code': None, 'venue': []}
                else:
                    # is an elective course
                    SCList = [s for s in sc.split('/') if s]
                    # strip trailing L/P/T
                    SCList = [s[:-1] if s[-1] in ('L', 'P', 'T') else s for s in SCList]

                    try:
                        VList = item['venue'].split('/')
                    except Exception:
                        VList = [None for _ in SCList]

                    try:
                        TCList = item['teacherCode'].split('/')
                    except Exception:
                        TCList = [None for _ in SCList]

                    if len(SCList) != len(VList):
                        VList = [None for _ in SCList]
                    if len(SCList) != len(TCList):
                        TCList = [None for _ in SCList]

                    for i in range(len(SCList)):
                        myList2.setdefault(SCList[i], []).append(subgroup)
                        subgroupDict['elective'][SCList[i]] = {
                            'lecture': {'slots': [], 'teacher_code': None, 'venue': []},
                            'lab': {'slots': [], 'teacher_code': None, 'venue': []},
                            'tutorial': {'slots': [], 'teacher_code': None, 'venue': []}
                        }

            # second pass: populate slots and teacher/venue
            for item in tt:
                sc = item.get('subjectCode', '').strip()
                if '/' not in sc and sc[:6] not in ['UMC743', 'UCS658', 'UCS657', 'UCS748', 'UCS539']:
                    tempDict = subgroupDict[item['type']][item['subjectCode']]
                    tempDict['slots'].append((item['t1'], item['t2']))
                    tempDict['teacher_code'] = item['teacherCode']
                    tempDict['venue'].append(item['venue'])
                else:
                    SCList = [s for s in sc.split('/') if s]
                    SCList = [s[:-1] if s[-1] in ('L', 'P', 'T') else s for s in SCList]

                    try:
                        VList = item['venue'].split('/')
                    except Exception:
                        VList = [None for _ in SCList]

                    try:
                        TCList = item['teacherCode'].split('/')
                    except Exception:
                        TCList = [None for _ in SCList]

                    if len(SCList) != len(VList):
                        VList = [None for _ in SCList]
                    if len(SCList) != len(TCList):
                        TCList = [None for _ in SCList]

                    for i in range(len(SCList)):
                        try:
                            tempDict2 = subgroupDict['elective'][SCList[i]][item['type']]
                            tempDict2['slots'].append((item['t1'], item['t2']))
                            tempDict2['teacher_code'] = TCList[i]
                            tempDict2['venue'].append(VList[i])
                        except Exception:
                            print(f"Error Occurred for {subgroup}")

        return myList, myList2

    @staticmethod
    def process_workbook(workbook):
        finalDict = {}
        finalDict2 = {}

        sheetnamesTemp = workbook.sheetnames
        sheetnames = []
        for s in sheetnamesTemp:
            s1 = s.strip()
            if s1 and s1[0] in ['1', '2', '3', '4']:
                sheetnames.append(s)

        for sheetname in sheetnames:
            first_char = sheetname.strip()[0]
            if first_char in ('1', '2', '3', '4'):
                sheet = workbook[sheetname]
                mr = sheet.merged_cells.ranges
                subgroupMap = PreprocessClass.create_subgroup_map(sheet)

                hrINC = 2
                if first_char == '4':
                    hrINC = 1

                tempDict, tempDict2 = PreprocessClass.get_time_table2(subgroupMap, sheet, mr, hrINC)

                for subg in tempDict.keys():
                    print(subg)
                    finalDict[subg] = tempDict[subg]

                for subcode in tempDict2.keys():
                    if subcode in finalDict2.keys():
                        finalDict2[subcode] = tempDict2[subcode] + finalDict2[subcode]
                    else:
                        finalDict2[subcode] = tempDict2[subcode]

        return finalDict, finalDict2

